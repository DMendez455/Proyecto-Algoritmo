from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib 
import pandas as pd

root = Tk()
root.title("Doctor Mario")
root.geometry('800x500+300+200')
root.resizable(False, False)
root.configure(bg="#326273")

pathfile = 'Backend_data.xlsx'
sheetName = 'Sheet'
header = 0

# Creamos los enlaces a Excel, identificamos las columnas que llevarán cada campo
file = pathlib.Path('Backend_data.xlsx')
if not file.exists():
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "nombre completo"
    sheet['B1'] = "sexo"
    sheet['C1'] = "edad"
    sheet['D1'] = "Direccion"
    sheet['F1'] = "fecha"
    sheet['E1'] = "Numero de celular"
    sheet['G1'] = "hora" 
    file.save('Backend_data.xlsx')
    
def AbrirArchivoExcel():
    df = pd.read_excel('Backend_data.xlsx',sheet_name='Sheet', header=0)
    return df

def FiltrarColumna(nombreColumna):
    df = AbrirArchivoExcel()
     # Se ecaonar una columna
    print("\nselección de la Columna: " + nombreColumna)
    col = df[nombreColumna]
    print(col)
    return col

def ValidarFechaCita(fechaAValidar):
    resultado = False
    df = AbrirArchivoExcel()
    col = FiltrarColumna('fecha')

    resultadoFiltro = df[col == fechaAValidar]
    # print(resultadoFiltro)
    # print('cantidad de registron encontrados es: ' + str(len(resultadoFiltro)))
    #print(df.head())
    resultado = len(resultadoFiltro) < 8
    #print('la fecha consultada ' +fechaAValidar + ' es:_'+ str(resultado)+'_')
    return resultado

def ValidarHora(horaCita, fechaAValidar):
    resultado = False
    df = AbrirArchivoExcel()

    resultadoFiltro = df[(df['fecha'] == fechaAValidar) & (df['hora'] == horaCita)]
    print(resultadoFiltro)
    print('cantidad de registron encontrados es: ' + str(len(resultadoFiltro)))
    resultado = len(resultadoFiltro) < 1
    return resultado

def submit():
    nombre = nombreValue.get()
    sexo = sexo_combobox.get()
    edad = edadValue.get()
    direccion = direccionEntry.get(1.0, END)
    contacto = contactoValue.get()
    fecha = fechaValue.get()
    hora = hora_combobox.get()


    # Validación de campos
    if not (nombre and sexo and direccion and fecha):
        messagebox.showerror("Campos Incompletos", "Por favor, ingrese toda la información obligatoria.")
        return
    # Validación de campos numéricos
    if not edad.isdigit():
        messagebox.showerror("Error en Edad", "La edad debe ser un número entero.")
        return
    if not contacto.isdigit():
        messagebox.showerror("Error", "los datos del contacto deben ser un numero de 8 digitos.")
        return
    if not ValidarFechaCita(fecha):
        messagebox.showerror("Error", "No existe cupo para la fecha seleccionada")
        return
    
    if not ValidarHora(hora,fecha):
        messagebox.showerror("Error", "No existe cupo para la hora seleccionada")
        return

    file = openpyxl.load_workbook(pathfile)
    sheet = file.active
    # Lugar de destino de los datos en la hoja de Excel
    sheet.cell(column=1, row=sheet.max_row + 1, value=nombre)
    sheet.cell(column=2, row=sheet.max_row, value=sexo)
    sheet.cell(column=3, row=sheet.max_row, value=edad)
    sheet.cell(column=4, row=sheet.max_row, value=direccion)
    sheet.cell(column=5, row=sheet.max_row, value=contacto)
    sheet.cell(column=6, row=sheet.max_row, value=fecha)
    sheet.cell(column=7, row=sheet.max_row, value=hora)

    # Guardamos en el Excel creado
    file.save(pathfile)


    messagebox.showinfo("guardado", "cita guardada con exito.")
    #clear()



def clear():
    nombreValue.set('')
    edadValue.set('')
    direccionEntry.delete(1.0, END)
    fechaValue.set('')
    contactoValue.set('')
    contactoEntry.delete(1.0, END)
    horaValue.set('') 

def abrir():
    fecha = fechaValue.get()
    df = AbrirArchivoExcel()
    col = FiltrarColumna('fecha')
    resultadoFiltro = df[col == fecha]

    # Crear una nueva ventana para mostrar la lista de citas
    segunda_ventana = tk.Toplevel(root)
    segunda_ventana.title("Citas para la fecha " + fecha)
    segunda_ventana.geometry('600x400+400+200')
    segunda_ventana.resizable(False, False)
    segunda_ventana.configure(bg="#326273")

    # Crear un Text Widget en la segunda ventana para mostrar las citas
    citas_text = Text(segunda_ventana, wrap=WORD, width=50, height=20, font=("Arial", 12))
    citas_text.pack()

    # Mostrar las citas en el Text Widget
    citas_text.insert(END, "Citas para la fecha " + fecha + ":\n\n")
    if not resultadoFiltro.empty:
        for index, row in resultadoFiltro.iterrows():
            citas_text.insert(END, f"Nombre: {row['nombre completo']}\n")
            citas_text.insert(END, f"Sexo: {row['sexo']}\n")
            citas_text.insert(END, f"Edad: {row['edad']}\n")
            citas_text.insert(END, f"Dirección: {row['Direccion']}\n")
            citas_text.insert(END, f"Contacto: {row['Numero de celular']}\n")
            citas_text.insert(END, f"Hora: {row['hora']}\n\n")
    else:
        citas_text.insert(END, "No hay citas para esta fecha.")

# ...

Button(root, text="Revisar Cita", bg="#326273", fg="white", width=15, height=2, command=abrir).place(x=550, y=350)

    
    # # cargamos el archivo excel
    #archivo_exel="Backend_data.xlsx"
    #libro_trabajo = openpyxl.load_workbook(archivo_exel)

    # accedemos a la hoja de trabajo
    #hoja=libro_trabajo["Sheet"]

    # #Recorremos las filas de la hoja
    #for "fila" in hoja.iter_rows(min_row=1, values_only=True):
    #print("fila")
    
    # #recorremos las columnas de la hoja
    #for columna in hoja.iter_cols(values_only=True):
    #print(columna)
    # #recorremos las celdas de la hoja de calculo
    #for fila in hoja.iter_rows(min_row=1, values_only=True):
    #for celda in fila:
    #print(celda)
    # #cerramos el archivo para optimizar recursos
    #libro_trabajo.close()

    #segunda_ventana = tk.Toplevel(root)
    #segunda_ventana.title("Citas por día")
    #segunda_ventana.geometry('800x400+400+300')
    #segunda_ventana.resizable(False, False)
    #segunda_ventana.configure(bg="#326273")


    

# Icono
icon_imagen = PhotoImage(file="DM.png")
root.iconphoto(False, icon_imagen)

# Título
Label(root, text="Por favor llenar el formulario:", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Etiquetas
Label(root, text='nombre', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='sexo.', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='edad', font=23, bg="#326273", fg="#fff").place(x=350, y=150)
Label(root, text='direccion', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='fecha', font=10, bg="#326273", fg="#fff").place(x=51, y=250)
Label(root, text='hora', font=23, bg="#326273", fg="#fff").place(x=350, y=250)
Label(root, text='Contacto', font=23, bg="#326273", fg="#fff").place(x=50, y=300)

# Entradas
nombreValue = StringVar()
edadValue = StringVar()
direccionValue = StringVar()
fechaValue = StringVar()
contactoValue = StringVar()
horaValue = StringVar()

nombreEntry = Entry(root, textvariable=nombreValue, width=45, bd=2, font=20)
edadEntry = Entry(root, textvariable=edadValue, width=21, bd=2, font=20)
direccionEntry = Text(root, width=45, height=2, bd=3.5)
fechaEntry = Entry(root, textvariable=fechaValue, width=15, bd=2, font=20)
contactoEntry = Entry(root, textvariable=contactoValue, width=45, bd=2, font=20)
horaEntry = Entry(root, textvariable=horaValue, width=21, bd=2, font=20)

nombreEntry.place(x=150, y=100)
edadEntry.place(x=410, y=150)
direccionEntry.place(x=150, y=200)
fechaEntry.place(x=150, y=250)
contactoEntry.place(x=150, y=300)
horaEntry.place()

# Género del paciente
sexo_combobox = Combobox(root, values=['MASCULINO', 'FEMENINO'], font='arial 14', state='r', width=14)
sexo_combobox.place(x=150, y=150)
sexo_combobox.set('FEMENINO')

# hora del paciente
hora_combobox = Combobox(root, values=['8:00 am', '9:00 am' , '10:00 am' , '11:00 am', '12:00 am', '2:00 pm', '3:00 pm', '4:00 pm'], font='arial 14', state='r', width=14)
hora_combobox.place(x=410, y=250)
hora_combobox.set('8:00 am')




Button(root, text="Guardar", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=100, y=350)
Button(root, text="Limpiar", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=250, y=350)
Button(root, text="Salir", bg="#326273", fg="white", width=15, height=2, command=lambda: root.destroy()).place(x=400, y=350)
Button(root, text="Revisar Cita", bg="#326273", fg="white", width=15, height=2, command=abrir).place(x=550, y=350)

root.mainloop()


